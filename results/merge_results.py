#!/usr/bin/env python3
"""Merge per-solver CSV results into a single XLSX workbook.

Read-only against the source CSVs; nothing in the input directory is modified.

Conventions
-----------
* Solver order is configurable. By default it follows:
      dlx, dxz, dxd-t1, dxd-t8, dyndxd-t1, dyndxd-t8
  Each entry maps to a CSV file via SOLVER_FILES below. Drop a new CSV into the
  results directory and append its (display name, filename) pair to extend.
* Instance names are stripped of the trailing ``.txt`` extension.
* Solution counts with > 11 digits are rendered in scientific notation
  (e.g. 2.0064e+22) so the cell stays readable but the underlying value is
  preserved as a string for big integers that exceed Excel's float range.
* A row is flagged in the ``All TO`` column when *every* present solver timed
  out on that instance.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Display name -> CSV filename. Order here is the column order in the workbook.
# Append new (name, filename) pairs to integrate additional solvers.
SOLVER_FILES: List[Tuple[str, str]] = [
    ("DLX",         "DLX_results.csv"),
    ("DXZ",         "DXZ_results.csv"),
    ("DXD-T1",      "DXD_t1_results.csv"),
    ("DXD-T8",      "DXD_t8_results.csv"),
    ("DynDXD-T1",   "DynDXD_t1_results.csv"),
    ("DynDXD-T8",   "DynDXD_t8_results.csv"),
]

TIMEOUT_TOKEN = "TO"
WARNING_RE = re.compile(r"\s*\(WARNING:[^)]*\)\s*$")


def strip_warning(text: str) -> str:
    """Remove a trailing ``(WARNING: ...)`` annotation if present."""
    return WARNING_RE.sub("", text).strip()


def is_timeout(time_cell: str) -> bool:
    return strip_warning(time_cell).upper() == TIMEOUT_TOKEN


def parse_time(time_cell: str) -> Optional[float]:
    cleaned = strip_warning(time_cell)
    if not cleaned or cleaned.upper() == TIMEOUT_TOKEN:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def format_solutions(raw: str) -> Tuple[Optional[object], str]:
    """Return (excel_value, display_text).

    For solution counts whose decimal representation has more than 11 digits we
    fall back to a power of 10 notation as a string, since Excel cannot store
    integers wider than 15 significant digits without loss.
    """
    raw = (raw or "").strip()
    if not raw:
        return None, ""

    digits = re.sub(r"\D", "", raw)
    if not digits or not raw.lstrip("-").isdigit():
        # Non-numeric (warnings, blanks, errors): keep as text.
        return raw, raw

    if len(digits) > 11:
        # Render in power of 10 notation: 2.0064 × 10^22 style.
        n = int(raw)
        sign = "-" if n < 0 else ""
        s = digits  # already only digits

        # 保留4位小数的尾数
        mantissa = f"{int(s[0])}.{s[1:5]}" if len(s) > 1 else s
        exp = len(s) - 1

        # 改写为 10 的次方格式
        power_str = f"{sign}{mantissa} × 10^{exp}"
        return power_str, power_str

    # Fits comfortably as an integer.
    return int(raw), raw


def load_solver_csv(path: Path) -> Dict[str, Dict[str, str]]:
    """Return ``{instance_basename: {field: value}}`` from a solver CSV."""
    out: Dict[str, Dict[str, str]] = {}
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return out
        # We rely on the column position rather than the (per-solver) name of
        # the time column, because the header for the time column varies
        # (DLX, DXZ, DXD_T1, ...).
        # Layout: Instance,#cols,#rows,<TIME>,Solutions,Max Blocks,|ZDD|,|DNNF|
        for row in reader:
            if not row or not row[0].strip():
                continue
            inst = row[0].strip()
            inst_key = inst[:-4] if inst.lower().endswith(".txt") else inst
            def cell(i: int) -> str:
                return row[i].strip() if i < len(row) else ""
            out[inst_key] = {
                "cols":      cell(1),
                "rows":      cell(2),
                "time":      cell(3),
                "solutions": cell(4),
            }
    return out


def merge(input_dir: Path, output: Path, solvers: List[Tuple[str, str]]) -> None:
    # Load each solver's CSV that actually exists.
    loaded: List[Tuple[str, Dict[str, Dict[str, str]]]] = []
    missing: List[str] = []
    for name, fname in solvers:
        csv_path = input_dir / fname
        if not csv_path.exists():
            missing.append(fname)
            continue
        loaded.append((name, load_solver_csv(csv_path)))
    if missing:
        print(f"[warn] missing CSVs (skipped): {', '.join(missing)}", file=sys.stderr)
    if not loaded:
        raise SystemExit(f"no solver CSVs found under {input_dir}")

    # Union of all instances, sorted alphabetically (case-insensitive).
    all_instances = sorted(
        {inst for _, data in loaded for inst in data.keys()},
        key=lambda s: s.lower(),
    )

    # First-seen #cols/#rows wins (skipping blanks).
    inst_dims: Dict[str, Tuple[str, str]] = {}
    for inst in all_instances:
        c = r = ""
        for _, data in loaded:
            row = data.get(inst)
            if row and (row["cols"] or row["rows"]):
                c, r = row["cols"], row["rows"]
                break
        inst_dims[inst] = (c, r)

    # First-seen non-empty solutions value across solvers.
    inst_solutions: Dict[str, str] = {}
    for inst in all_instances:
        sol = ""
        for _, data in loaded:
            row = data.get(inst)
            if row and row["solutions"]:
                sol = row["solutions"]
                break
        inst_solutions[inst] = sol

    # Build workbook.
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    header = ["Instance", "#cols", "#rows"] + [name for name, _ in loaded] + ["Solutions", "All TO"]
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    timeout_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
    all_to_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for inst in all_instances:
        cols, rows_ = inst_dims[inst]
        row_values: List[object] = [inst, _maybe_int(cols), _maybe_int(rows_)]
        timeout_flags: List[bool] = []
        for name, data in loaded:
            row = data.get(inst)
            if row is None:
                row_values.append("")
                timeout_flags.append(False)  # absent != timed-out
                continue
            if is_timeout(row["time"]):
                row_values.append("TO")
                timeout_flags.append(True)
            else:
                t = parse_time(row["time"])
                row_values.append(t if t is not None else row["time"])
                timeout_flags.append(False)
        sol_value, _ = format_solutions(inst_solutions[inst])
        row_values.append(sol_value if sol_value is not None else "")
        # All-TO: every solver that has an entry on this instance timed out,
        # and at least one solver actually ran it.
        present_flags = [
            (data.get(inst) is not None, t) for (_, data), t in zip(loaded, timeout_flags)
        ]
        present_to = [t for p, t in present_flags if p]
        all_to = bool(present_to) and all(present_to)
        row_values.append("YES" if all_to else "")
        # Replace None placeholders with empty strings for cleaner display.
        row_values = [("" if v is None else v) for v in row_values]
        ws.append(row_values)

        excel_row = ws.max_row
        # Highlight individual TO cells.
        for offset, flag in enumerate(timeout_flags):
            if flag:
                ws.cell(row=excel_row, column=4 + offset).fill = timeout_fill
        if all_to:
            ws.cell(row=excel_row, column=len(header)).fill = all_to_fill

    # Format time columns: 4 decimal places when numeric.
    time_col_start = 4
    time_col_end = 3 + len(loaded)
    for col in range(time_col_start, time_col_end + 1):
        for row in range(2, ws.max_row + 1):
            c = ws.cell(row=row, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "0.0000"

    # Auto-size columns (rough heuristic).
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 8), min(40, len(v) + 2))
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "B2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"wrote {output} ({ws.max_row - 1} instances, {len(loaded)} solvers)")


def _maybe_int(s: str):
    s = (s or "").strip()
    if not s:
        return ""
    try:
        return int(s)
    except ValueError:
        return s


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "input_dir",
        nargs="?",
        default="batch_results_20260511",
        help="directory containing per-solver *_results.csv files",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="output XLSX path (default: <input_dir>/merged_results.xlsx)",
    )
    args = parser.parse_args()
    in_dir = Path(args.input_dir).resolve()
    out = Path(args.output) if args.output else in_dir / "merged_results.xlsx"
    merge(in_dir, out, SOLVER_FILES)


if __name__ == "__main__":
    main()
